import openpyxl
from data_helper import  char_tokenizer,load_tokenizer
from charcnn import CHARCNN
import numpy as np
import os
from metrics import *
import matplotlib.pyplot as plt
import config
from keras.utils import to_categorical

FIELD_SLICE_DICT = {"装运": "PORT",
             "卸货": "PORT",
             "船名": "VESSEL",
             "合同": "NO",
             "LC": "LCNO",
             "BL": "BLNO",
             "发票": "INNO"}


def type_one_hot(slice_type):
    type_idx = config.FIELD_TYPE_DICT[slice_type]
    if slice_type in config.FIELD_TYPE_DICT:
        slice_type_one_hot = to_categorical(type_idx, len(config.FIELD_TYPE_DICT)+1)
    else:
        slice_type_one_hot = to_categorical(type_idx, len(config.FIELD_TYPE_DICT))
    return slice_type_one_hot


def predict_excel(excel_path, pre_model):

    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    sheet.title = "test_res"

    sheet.cell(row=1, column=8, value="预测概率")
    sheet.cell(row=1, column=9, value="是否为目标切片")
    sheet.cell(row=1, column=10, value="分类概率")

    pred_rate = []
    cur_type = ""
    cur_region_idx = -1
    field_ocr = ""
    tokenizer = load_tokenizer('./tokenizer/tokenizer.pickle')
    for i in range(1, sheet.max_row):
        region_idx = str(sheet.cell(row=i + 1, column=1).value)

        if str(cur_region_idx) == region_idx and cur_type:
            field_type = cur_type
        else:
            field_type = sheet.cell(row=i + 1, column=2).value
            cur_type = field_type
            cur_region_idx = region_idx

        loc_info = sheet.cell(row=i + 1, column=5).value
        ocr_info = sheet.cell(row=i + 1, column=6).value
        flag_info = sheet.cell(row=i + 1, column=7).value
        print(region_idx, field_type, loc_info, ocr_info, flag_info)

        if flag_info and int(flag_info) == 0:
            field_ocr = ocr_info
            continue

        print(region_idx, field_type, loc_info, ocr_info)

        if field_type and loc_info and ocr_info:
            loc_info = str(loc_info)
            ocr_info = str(ocr_info)
            #print(region_idx, field_type, loc_info, ocr_info)
            locs = [float(loc) for loc in loc_info[1:-1].replace(" ", "").split(',')]
            field_type_start = str(field_type)[:2]
            field_slice_type = FIELD_SLICE_DICT[field_type_start]

            slice_type_one_hot = type_one_hot(field_type_start)

            #_char = "$" + field_ocr + '\t' + ocr_info + "\n"
            _char = "$" + field_slice_type + '\t' + ocr_info + "\n"


            char_idx_pad = char_tokenizer([_char], tokenizer=tokenizer)

            locs = list(locs) + list(slice_type_one_hot)
            print(_char, locs)
            res = pre_model.predict_one(np.array(char_idx_pad), np.array([locs]))
            print(res[0], res[1])

            sheet.cell(row=i + 1, column=8, value=str(res[1][0][0]))
            pred_rate.append(res[1][0][0])

            if res[1][0][0] >= 0.5:
                sheet.cell(row=i + 1, column=9, value="目标值")
            sheet.cell(row=i + 1, column=10, value=str(res[0][0][0]))
    # for j in range(0, 20):
    #     sheet.cell(row=j+2, column=8, value=str(1))

    workbook.save(excel_path.replace('result', 'result_test'))
    print("xlsx格式表格写入数据成功！")
    return pred_rate


if __name__ == "__main__":

    #excel_dir = '/Users/peng_ji/Desktop/AB_test/'
    excel_dir = '/Users/peng_ji/codeHub/rookieCode/AB_test/'
    trans_id = os.listdir(excel_dir)
    print(trans_id)
    all_data = []
    charcnn = CHARCNN()
    charcnn.charcnn_merge_loc_model()

    #charcnn.load_weights('./saved_model/model-dense080405-512-1-5-st-weights-12-0.97.hdf5')
    charcnn.load_weights('./saved_model/model-dense08-512-1-5-st4d-neg-weights-05-0.98.hdf5') #目前最好

    pred_rate_all = []

    for id in trans_id:
        trans_dir = os.path.join(excel_dir, id)

        if os.path.isdir(trans_dir):
            excel_path = os.path.join(trans_dir, 'result.xlsx')
            print(excel_path)
            pred_rate= predict_excel(excel_path, charcnn)
            pred_rate_all.extend(pred_rate)