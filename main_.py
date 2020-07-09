import openpyxl
from data_helper import  char_tokenizer, load_tokenizer
from bilstm import BiLSTM
import numpy as np
import os
from metrics import *
import matplotlib.pyplot as plt
import pandas as pd

FIELD_SLICE_DICT = {"装运": "PORT",
             "卸货": "PORT",
             "船名": "VESSEL",
             "合同": "NO",
             "LC": "LCNO",
             "BL": "BLNO",
             "发票": "INNO"}


def addRes4NoRes(res_dict):
    res_df = pd.DataFrame(res_dict)
    region_df = res_df.groupby(["region_idx"])
    row_idxs = []
    for name, rdf in region_df:
        # print("===========================")
        # print(name)
        # print(rdf)
        if rdf["res"].sum() == 0:
            maxres_i = rdf['res_prob'].idxmax()
            maxres_class_prob = rdf.loc[maxres_i, 'class_prob']
            maxres_row_idx = rdf.loc[maxres_i, 'row_idx']
            if maxres_class_prob > 0.5:
                row_idxs.append([maxres_i, maxres_row_idx])
    return row_idxs


def predict_excel(excel_path, pre_model):
    field_res_dict = {"装运": [[],[]],
             "卸货": [[],[]],
             "船名": [[],[]],
             "合同": [[],[]],
             "LC": [[],[]],
             "BL": [[],[]],
             "发票": [[],[]]}

    # 用于在无目标区域内寻找合理的目标
    res_dict = {"region_idx":[], "class_prob":[],"res_prob":[],"row_idx":[], "res":[]}

    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    sheet.title = "test_res"


    sheet.cell(row=1, column=8, value="预测概率")
    sheet.cell(row=1, column=9, value="是否为目标切片")
    sheet.cell(row=1, column=10, value="分类概率")
    true_label = []
    pred_rate = []
    tokenizer = load_tokenizer('./tokenizer/tokenizer.pickle')
    for i in range(1, sheet.max_row):
        region_idx = str(sheet.cell(row=i + 1, column=1).value)
        field_type = sheet.cell(row=i + 1, column=2).value
        slice_type = sheet.cell(row=i + 1, column=3).value
        loc_info = sheet.cell(row=i + 1, column=5).value
        ocr_info = sheet.cell(row=i + 1, column=6).value

        if slice_type and field_type and loc_info and ocr_info:
            loc_info = str(loc_info)
            ocr_info = str(ocr_info)
            print(region_idx, field_type, slice_type, loc_info, ocr_info)
            locs = [float(loc) for loc in loc_info[1:-1].replace(" ", "").split(',')]
            field_type_start = str(field_type)[:2]
            field_slice_type = FIELD_SLICE_DICT[field_type_start]

            if "其他" in field_type:
                true_label.append(0)
                field_res_dict[field_type_start][0].append(0)
            else:
                true_label.append(1)
                field_res_dict[field_type_start][0].append(1)


            _char = "$" + field_slice_type + "\t" + ocr_info + "\n"

            char_idx_pad = char_tokenizer([_char], tokenizer=tokenizer)
            #char_idx_pad = char_tokenizer([_char], './tokenizer/tokenizer_nopunc.pickle')

            res = pre_model.predict_one(np.array(char_idx_pad), np.array([locs]))
            res_dict["region_idx"].append(region_idx)
            res_dict["class_prob"].append(res[0][0][0])
            res_dict["res_prob"].append(res[1][0][0])
            res_dict["row_idx"].append(i)
            print(res[0], res[1])

            sheet.cell(row=i + 1, column=8, value=str(res[1][0][0]))
            pred_rate.append(res[1][0][0])
            field_res_dict[field_type_start][1].append(res[1][0][0])
            if res[1][0][0] >= 0.5:
                sheet.cell(row=i + 1, column=9, value="目标值")
                res_dict["res"].append(1)
            else:
                res_dict["res"].append(0)
            sheet.cell(row=i + 1, column=10, value=str(res[0][0][0]))
    row_idxs = addRes4NoRes(res_dict)
    for i, ri in row_idxs:
        sheet.cell(row=ri + 1, column=9, value="后增目标值")
        # 由于改过来了所以认为识别为真
        pred_rate[i] = 1.0

    workbook.save(excel_path.replace('result', 'result_test'))
    print("xlsx格式表格写入数据成功！")
    return true_label, pred_rate, field_res_dict

if __name__ == "__main__":

    excel_dir = '/Users/peng_ji/Desktop/AB_train/'
    trans_id = os.listdir(excel_dir)
    #trans_id = ['732601AB20000141']
    print(trans_id)
    all_data = []
    bilstm = BiLSTM()
    bilstm.charcnn_merge_loc_model()
    bilstm.load_weights('./saved_model/model-weights-63-0.96.hdf5')
    true_label_all = []
    pred_rate_all = []
    field_res_all_dict = {"装运": [[], []],
                      "卸货": [[], []],
                      "船名": [[], []],
                      "合同": [[], []],
                      "LC": [[], []],
                      "BL": [[], []],
                      "发票": [[], []]}
    for id in trans_id:
        trans_dir = os.path.join(excel_dir, id)

        if os.path.isdir(trans_dir):
            excel_path = os.path.join(trans_dir, 'result.xlsx')
            print(excel_path)
            true_label, pred_rate, field_res_dict= predict_excel(excel_path, bilstm)
            true_label_all.extend(true_label)
            pred_rate_all.extend(pred_rate)
            for key in field_res_all_dict.keys():
                field_res_all_dict[key][0].extend(field_res_dict[key][0])
                field_res_all_dict[key][1].extend(field_res_dict[key][1])


    accs = []
    aucs = []
    precisions = []
    recalls = []
    metrics_dict = {"装运": [],
                      "卸货": [],
                      "船名": [],
                      "合同": [],
                      "LC": [],
                      "BL": [],
                      "发票": []}
    x = list(np.arange(0.05,0.8,0.05))
    for r in x:
        for key in field_res_all_dict.keys():
            pred_label_all = pred_probability2label(field_res_all_dict[key][1], r)
            acc, auc, precision, recall = genMetrics(field_res_all_dict[key][0], pred_label_all)
            avg = (acc + auc + precision + recall)/4.0
            metrics_dict[key].append([r, acc, auc, precision, recall, avg])
        pred_label_all = pred_probability2label(pred_rate_all, r)
        acc, auc, precision, recall = genMetrics(true_label_all, pred_label_all)
        avg = (acc + auc + precision + recall) / 4.0
        accs.append(acc)
        aucs.append(auc)
        precisions.append(precision)
        recalls.append(recall)

        print(r, acc, auc, precision, recall, avg)
    plt.plot(x, accs, label="accuracy")
    plt.plot(x, aucs, label="auc")
    plt.plot(x, precisions, label="precision")
    plt.plot(x, recalls, label="recall")
    plt.legend(loc='best')
    plt.show()
    #
    #
    # for key in metrics_dict.keys():
    #     print(key, "===============================")
    #     for v in metrics_dict[key]:
    #         print(v)



