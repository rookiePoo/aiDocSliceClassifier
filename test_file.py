import openpyxl
from data_helper import  char_tokenizer
from bilstm import BiLSTM
import numpy as np
import os
from metrics import *
import matplotlib.pyplot as plt

FIELD_SLICE_DICT = {"装运": "PORT",
             "卸货": "PORT",
             "船名": "TXT",
             "合同": "NO",
             "LC": "LCNO",
             "BL": "BLNO",
             "发票": "INNO"}



def test(excel_path, pre_model):
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    sheet.title = "test_res"


    sheet.cell(row=1, column=8, value="预测概率")
    sheet.cell(row=1, column=9, value="是否为目标切片")
    true_label = []
    pred_rate = []
    for i in range(1, sheet.max_row):
        region_idx = str(sheet.cell(row=i + 1, column=1).value)
        field_type = sheet.cell(row=i + 1, column=2).value
        slice_type = sheet.cell(row=i + 1, column=3).value
        loc_info = sheet.cell(row=i + 1, column=5).value
        ocr_info = sheet.cell(row=i + 1, column=6).value

        if slice_type and field_type and loc_info and ocr_info:
            loc_info = str(loc_info)
            ocr_info = str(ocr_info)
            if "其他" in field_type:
                true_label.append(0)
            else:
                true_label.append(1)
            print(region_idx, field_type, slice_type, loc_info, ocr_info)
            locs = [float(loc) for loc in loc_info[1:-1].replace(" ", "").split(',')]
            field_type_start = str(field_type)[:2]
            field_slice_type = FIELD_SLICE_DICT[field_type_start]

            _char = "$" + field_slice_type + "\t" + ocr_info + "\n"
            char_idx_pad = char_tokenizer([_char], './tokenizer/tokenizer.pickle')

            res = pre_model.predict_one(np.array(char_idx_pad), np.array([locs]))
            print(res[1])

            sheet.cell(row=i + 1, column=8, value=str(res[1][0][0]))
            pred_rate.append(res[1][0][0])
            if res[1][0][0] >= 0.5:
                sheet.cell(row=i + 1, column=9, value="目标值")
    # for j in range(0, 20):
    #     sheet.cell(row=j+2, column=8, value=str(1))

    workbook.save(excel_path.replace('result', 'result_test'))
    print("xlsx格式表格写入数据成功！")
    return true_label, pred_rate

if __name__ == "__main__":

    excel_dir = '/Users/peng_ji/Desktop/labeled01/'
    trans_id = os.listdir(excel_dir)
    print(trans_id)
    all_data = []
    bilstm = BiLSTM()
    bilstm.charcnn_merge_loc_model()
    bilstm.load_weights('./saved_model/model-weights-55-0.96.hdf5')
    true_label_all = []
    pred_rate_all = []
    for id in trans_id:
        trans_dir = os.path.join(excel_dir, id)

        if os.path.isdir(trans_dir):
            excel_path = os.path.join(trans_dir, 'result.xlsx')
            print(excel_path)
            true_label, pred_rate= test(excel_path, bilstm)
            true_label_all.extend(true_label)
            pred_rate_all.extend(pred_rate)


    accs = []
    aucs = []
    precisions = []
    recalls = []
    x = list(np.arange(0.05,0.8,0.05))
    for r in x:
        pred_label_all = pred_probability2label(pred_rate_all, r)
        acc, auc, precision, recall = genMetrics(true_label_all, pred_label_all)
        accs.append(acc)
        aucs.append(auc)
        precisions.append(precision)
        recalls.append(recall)

        print(acc, auc, precision, recall)
    plt.plot(x, accs, label="accuracy")
    plt.plot(x, aucs, label="auc")
    plt.plot(x, precisions, label="precision")
    plt.plot(x, recalls, label="recall")
    plt.legend(loc='best')
    plt.show()



